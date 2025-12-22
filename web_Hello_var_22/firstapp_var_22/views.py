# firstapp_var_22/views.py
from pathlib import Path
import csv
from openpyxl import load_workbook
from decimal import Decimal
from datetime import datetime
from django import forms
from .forms import OrderForm, SimpleOrderForm
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Count, Sum, Avg
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from .models import Book, Category, User, Order, OrderItem

BASE_DIR = Path(__file__).resolve().parent
TABLICA_DIR = BASE_DIR / 'tablica'


# === Формы ===
class OrderForm(forms.Form):
    """Форма для оформления заказа"""
    quantity = forms.IntegerField(
        label='Количество',
        min_value=1,
        initial=1,
        widget=forms.NumberInput(attrs={'class': 'form-control'})
    )
    shipping_address = forms.CharField(
        label='Адрес доставки',
        widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        required=True
    )
    payment_method = forms.ChoiceField(
        label='Способ оплаты',
        choices=Order.PAYMENT_METHODS,
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    notes = forms.CharField(
        label='Примечания',
        widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        required=False
    )


class SimpleOrderForm(forms.Form):
    """Упрощенная форма для быстрой покупки"""
    quantity = forms.IntegerField(
        min_value=1,
        initial=1,
        widget=forms.NumberInput(attrs={'class': 'form-control'})
    )


# === Главные ===
def index(request):
    return render(request, 'index.html')


def navigator(request):
    return render(request, 'navigator.html')


def thank_you(request):
    return render(request, 'thank_you.html')

def fastapi_redirect(request):
    """Страница-редирект на FastAPI"""
    return render(request, 'fastapi_redirect.html', {
        'fastapi_url': 'http://localhost:8000',
        'docs_url': 'http://localhost:8000/docs'
    })

# === Книги ===
# firstapp_var_22/views.py - обновите функцию book_list
def book_list(request, page=1):
    per_page = 12
    books = Book.objects.all().select_related('category')

    # Фильтрация по категории - БЕЗОПАСНАЯ ОБРАБОТКА
    category_id = request.GET.get('category', '').strip()
    if category_id and category_id.isdigit():
        books = books.filter(category_id=int(category_id))

    # Фильтрация по наличию - БЕЗОПАСНАЯ ОБРАБОТКА
    available_filter = request.GET.get('available', '').strip()
    if available_filter == '1':
        books = books.filter(available=True)
    elif available_filter == '0':
        books = books.filter(available=False)
    # Если пусто или другое значение - не фильтруем

    # Сортировка с значением по умолчанию
    sort = request.GET.get('sort', '-created')
    valid_sorts = ['created', '-created', 'price', '-price', 'title']
    if sort in valid_sorts:
        books = books.order_by(sort)
    else:
        sort = '-created'  # Значение по умолчанию
        books = books.order_by(sort)

    # Пагинация
    paginator = Paginator(books, per_page)

    try:
        page_number = int(request.GET.get('page', page))
    except ValueError:
        page_number = 1

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Все категории для фильтра
    categories = Category.objects.all()

    return render(request, 'book_list.html', {
        'books': page_obj,
        'page_obj': page_obj,
        'categories': categories,
        'current_sort': sort,
        'current_category': category_id,
        'current_available': available_filter,
    })

def book_detail(request, slug, year=None):
    book = get_object_or_404(Book.objects.select_related('category'), slug=slug)
    return render(request, 'book_detail.html', {'book': book, 'year': year})


# === Оформление заказа ===
@login_required
def buy_book(request, book_id=None):
    """
    Оформление заказа. Требуется авторизация.
    Если book_id передан — покупка конкретной книги,
    иначе редирект на список книг.
    """
    if book_id:
        book = get_object_or_404(Book, pk=book_id, available=True)
    else:
        # Если книга не указана, показываем список книг
        return redirect('book_list')

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            quantity = data.get('quantity', 1)

            # Проверка наличия
            if book.stock < quantity:
                form.add_error('quantity', f'Доступно только {book.stock} экземпляров')
                return render(request, 'buy_book.html', {'form': form, 'book': book})

            # Рассчитываем сумму
            total_price = Decimal(book.price) * quantity

            # Создаем заказ
            order = Order.objects.create(
                user=request.user,
                payment_method=data.get('payment_method', 'CARD'),
                shipping_address=data.get('shipping_address', ''),
                notes=data.get('notes', ''),
                total=total_price
            )

            # Создаем позицию заказа
            OrderItem.objects.create(
                order=order,
                book=book,
                quantity=quantity,
                price=book.price
            )

            # Обновляем остаток
            book.stock -= quantity
            book.save(update_fields=['stock'])

            messages.success(request, f'Заказ #{order.id} успешно создан!')
            return redirect('order_success', pk=order.id)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        # Начальные данные формы
        initial = {
            'quantity': 1,
            'payment_method': 'CARD',
            'shipping_address': request.user.address if request.user.address else ''
        }
        form = OrderForm(initial=initial)

    return render(request, 'buy_book.html', {'form': form, 'book': book})


def order_success(request, pk):
    """Страница успешного оформления заказа"""
    order = get_object_or_404(Order.objects.select_related('user'), pk=pk)

    # Проверяем, что заказ принадлежит текущему пользователю
    if request.user.is_authenticated and order.user != request.user:
        messages.error(request, 'У вас нет доступа к этому заказу.')
        return redirect('index')

    return render(request, 'order_success.html', {'order': order})


def order_history(request):
    """История заказов пользователя"""
    if not request.user.is_authenticated:
        return redirect('login')

    orders = Order.objects.filter(user=request.user).order_by('-created')
    return render(request, 'order_history.html', {'orders': orders})


# === Tables / files ===
def tablica_display(request):
    rows = []
    csv_file = TABLICA_DIR / 'books.csv'
    if csv_file.exists():
        with open(csv_file, encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            rows = [row for row in reader]
    else:
        rows = [['Файл не найден', str(csv_file)]]
    return render(request, 'tablica_display.html', {'rows': rows, 'source': 'CSV'})


def tablica_from_text(request):
    rows = []
    txt_file = TABLICA_DIR / 'books.txt'
    if txt_file.exists():
        with open(txt_file, encoding='utf-8') as f:
            rows = [line.strip().split(';') for line in f if line.strip()]
    else:
        rows = [['Файл не найден', str(txt_file)]]
    return render(request, 'tablica_display.html', {'rows': rows, 'source': 'TXT'})


def tablica_from_xlsx(request):
    rows = []
    xlsx_file = TABLICA_DIR / 'books.xlsx'
    if xlsx_file.exists():
        wb = load_workbook(xlsx_file, read_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    else:
        rows = [['Файл не найден', str(xlsx_file)]]
    return render(request, 'tablica_display.html', {'rows': rows, 'source': 'XLSX'})


def products_demo(request):
    print("=" * 60)
    print("DEBUG: Функция products_demo вызвана")

    txt_rows, csv_rows, xlsx_rows = [], [], []

    # --- Проверка TXT ---
    txt_file = TABLICA_DIR / 'books.txt'
    print(f"TXT файл: {txt_file}")
    print(f"TXT существует: {txt_file.exists()}")

    if txt_file.exists():
        with open(txt_file, encoding='utf-8') as f:
            content = f.read()
            print(f"TXT содержимое: {content[:200]}...")
            f.seek(0)
            txt_rows = [line.strip().split(';') for line in f if line.strip()]
            print(f"Прочитано {len(txt_rows)} строк из TXT")
    else:
        print("ОШИБКА: TXT файл не найден!")

    # --- Проверка CSV ---
    csv_file = TABLICA_DIR / 'books.csv'
    print(f"CSV файл: {csv_file}")
    print(f"CSV существует: {csv_file.exists()}")

    if csv_file.exists():
        with open(csv_file, encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            csv_rows = [row for row in reader]
            print(f"Прочитано {len(csv_rows)} строк из CSV")
    else:
        print("ОШИБКА: CSV файл не найден!")

    # --- Проверка XLSX ---
    xlsx_file = TABLICA_DIR / 'books.xlsx'
    print(f"XLSX файл: {xlsx_file}")
    print(f"XLSX существует: {xlsx_file.exists()}")

    if xlsx_file.exists():
        try:
            wb = load_workbook(xlsx_file, read_only=True, data_only=True)
            ws = wb.active
            print(f"XLSX лист: {ws.title}")
            print(f"XLSX размеры: {ws.max_row} строк, {ws.max_column} столбцов")

            xlsx_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                xlsx_rows.append(list(row))
                if i <= 3:
                    print(f"  Строка {i}: {list(row)}")

            print(f"Прочитано {len(xlsx_rows)} строк из XLSX")
        except Exception as e:
            print(f"ОШИБКА при чтении XLSX: {e}")
    else:
        print("ОШИБКА: XLSX файл не найден!")

    print("=" * 60)

    return render(request, 'products_demo.html', {
        'txt_rows': txt_rows,
        'csv_rows': csv_rows,
        'xlsx_rows': xlsx_rows
    })


def redirect_old_books(request):
    return redirect('book_list')


def order_analysis(request):
    """Анализ заказов"""
    orders = Order.objects.all().select_related('user')

    # -------- ФИЛЬТРАЦИЯ --------
    user_id = request.GET.get('user')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if user_id:
        orders = orders.filter(user_id=user_id)

    if status:
        orders = orders.filter(status=status)

    if date_from:
        orders = orders.filter(created__date__gte=date_from)

    if date_to:
        orders = orders.filter(created__date__lte=date_to)

    # -------- СОРТИРОВКА --------
    sort = request.GET.get('sort', '-created')
    if sort in ['created', '-created', 'total', '-total']:
        orders = orders.order_by(sort)

    # -------- АГРЕГАТЫ --------
    stats = orders.aggregate(
        total_count=Count('id'),
        total_sum=Sum('total'),
        avg_sum=Avg('total')
    )

    users = User.objects.all()

    context = {
        'orders': orders,
        'users': users,
        'status_choices': Order.STATUS_CHOICES,
        'stats': stats,
    }
    return render(request, 'order_analysis.html', context)


def export_orders(request):
    """Экспорт заказов в CSV"""
    orders = Order.objects.all().select_related('user')

    # Применяем фильтры
    user_id = request.GET.get('user')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if user_id:
        orders = orders.filter(user_id=user_id)
    if status:
        orders = orders.filter(status=status)
    if date_from:
        orders = orders.filter(created__date__gte=date_from)
    if date_to:
        orders = orders.filter(created__date__lte=date_to)

    # Создаем CSV ответ
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="orders_report.csv"'},
    )

    writer = csv.writer(response)
    writer.writerow(['ID', 'Пользователь', 'Статус', 'Дата', 'Сумма', 'Способ оплаты'])

    for order in orders:
        writer.writerow([
            order.id,
            order.user.username,
            order.get_status_display(),
            order.created.strftime('%Y-%m-%d %H:%M'),
            order.total,
            order.get_payment_method_display()
        ])

    return response


def analysis(request):
    return render(request, 'analysis.html')


def reviews(request):
    # Временные данные для отзывов
    sample_reviews = [
        {'author': 'Иван', 'text': 'Очень хорошая книга!'},
        {'author': 'Мария', 'text': 'Доставка была быстрой, спасибо!'},
        {'author': 'Алексей', 'text': 'Большой выбор и отличное обслуживание.'},
    ]
    return render(request, 'reviews.html', {'reviews': sample_reviews})


def files_books(request):
    """Просмотр книг из файлов"""
    text_lines = []
    excel_books = []

    # --- Текстовый файл ---
    txt_file = TABLICA_DIR / 'books.txt'
    if txt_file.exists():
        with open(txt_file, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split(';')
                    if len(parts) >= 4:
                        text_lines.append(f"{parts[0]} — {parts[1]} — {parts[2]} ₽ — {parts[3]}")
    else:
        print("TXT файл не найден:", txt_file)

    # --- Excel файл ---
    xlsx_file = TABLICA_DIR / 'books.xlsx'
    if xlsx_file.exists():
        wb = load_workbook(xlsx_file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = rows[0]
            for row in rows[1:]:
                book_data = dict(zip(headers, row))
                excel_books.append({
                    'title': book_data.get('название', 'Нет названия'),
                    'author': book_data.get('автор', 'Нет автора'),
                    'price': book_data.get('цена', '0'),
                    'stock': book_data.get('продавец', 'Нет информации'),
                })
    else:
        print("XLSX файл не найден:", xlsx_file)

    context = {
        'text_lines': text_lines,
        'excel_books': excel_books,
    }
    return render(request, 'products_demo.html', context)